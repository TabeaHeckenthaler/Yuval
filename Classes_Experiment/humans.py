from trajectory import MatlabFolder
import scipy.io as sio
import numpy as np
from os import path, listdir
from openpyxl import load_workbook
import csv
from Classes_Experiment.forces import force_attachment_positions

excel_sheet_directory = '{0}{1}phys-guru-cs{2}ants{3}Tabea{4}Human Experiments'.format(path.sep, path.sep,
                                                                                       path.sep,
                                                                                       path.sep, path.sep)


def get_sheet():
    workbook = load_workbook(filename=excel_sheet_directory + path.sep + "Testable.xlsx")
    sheet = workbook.active
    return sheet


sheet = get_sheet()
number_exp = [i for i in range(1, int(sheet.dimensions.split(':')[1][1:]))
              if sheet.cell(row=i, column=1).value is not None][-1]

participant_number = {'Small Near': 1, 'Small Far': 1, 'Medium': 9, 'Large': 26}


def date(human):
    day = sheet.cell(row=excel_worksheet_index(human.filename), column=2).value
    daytime = sheet.cell(row=excel_worksheet_index(human.filename), column=6).value
    return day.replace(hour=daytime.hour, minute=daytime.minute)


def force_directory(human):
    day_string = str(date(human).year) + '-' + str(date(human).month) + '-' + str(date(human).day)
    return ('{0}{1}phys-guru-cs{2}ants{3}Tabea{4}Human Experiments{5}Raw Data and Videos{7}'
            + day_string + '{8}Force Measurements{9}' + human.size).format(path.sep, path.sep, path.sep, path.sep,
                                                                           path.sep, path.sep, path.sep, path.sep,
                                                                           path.sep, path.sep)


def force_filename(human):
    differences = [int(filename[:4]) - int(str(date(human).hour).zfill(2) + str(date(human).minute).zfill(2))
                   for filename in listdir(force_directory(human))]
    index = int(np.argmin(np.abs(differences)))
    text_file_name = listdir(force_directory(human))[index]
    return text_file_name


def convert_to_frames(fps, times):
    seconds = [int(time[0].split(':')[0]) * 3600 + int(time[0].split(':')[1]) * 60 + int(time[0].split(':')[2]) for time
               in times]
    seconds = [sec - seconds[0] for sec in seconds]
    frames = []
    for second in range(seconds[-1]):
        measurements_per_second = len([sec for sec in seconds if sec == second])
        for ii in range(measurements_per_second):
            frames.append(second * fps + int(ii * fps / measurements_per_second))
    return frames


def synchronization_offset(self, x):
    """ frame of turning on force meter relative to start of the raw movie """
    [minute, second] = [int(number) for number in
                        sheet.cell(row=excel_worksheet_index(self.filename), column=16).value.strip()[:-3].split(':')]
    frame_force_meter = (second + minute * 60) * x.fps

    """ if the frame of synchronization is BEFORE the start of the movie which was tracked """
    if sheet.cell(row=excel_worksheet_index(self.filename), column=16).value[0] == '-':
        frame_force_meter = - frame_force_meter

    """ time of tracking relative to start of the raw movie """
    frame_tracking = int(sheet.cell(row=excel_worksheet_index(self.filename), column=8).value)
    return frame_tracking - frame_force_meter


def relative_to_minimum(forces):
    forces = np.array(forces)
    for i in range(len(forces[0])):
        forces[:, i] = np.array(forces)[:, i] - min(forces[:, i])
    return forces


def force_debugger(human, forces_all_frames, x):
    if np.isnan(np.sum([human.frames[i].forces[1] for i in range(0, len(human.frames))])):
        with open(excel_sheet_directory + path.sep + 'mistakes.txt', 'a') as f:
            f.write('\n' + x.filename + '\n')
            f.write('original movie: ' + sheet.cell(row=excel_worksheet_index(human.filename), column=1).value + '\n')
            f.write('force file: ' + force_filename(human) + '\n')
            f.write(
                'configuration time: ' + sheet.cell(row=excel_worksheet_index(human.filename), column=16).value + '\n')
            f.write('length of force measurement: ' +
                    str(int(np.floor(len(forces_all_frames) / x.fps / 60))).zfill(2) + ':' +
                    str(int(np.floor(len(forces_all_frames) / x.fps % 60))).zfill(2) + '\n')
            f.write('missing frames: ' +
                    str(len([i for i in range(len(human.frames))
                             if np.isnan(human.frames[i].forces[0])])) + '\n')


class Humans_Frame:
    def __init__(self, size):
        self.position = np.zeros((participant_number[size], 2))
        self.angle = np.zeros(participant_number[size])
        self.carrying = np.zeros((participant_number[size]))
        self.major_axis_length = list()
        # self.forces = list()


class Humans:
    def __init__(self, x):
        self.filename = x.filename
        self.VideoChain = [x.filename]
        self.frames = list()
        self.size = x.size
        self.number = len(self.gender())
        self.occupied = list(self.gender().keys())
        if isinstance(sheet.cell(row=excel_worksheet_index(self.filename), column=16).value, str):
            self.matlab_human_loading(x)
            self.forces_loading(x)
        self.gender_string = self.gender()

    def matlab_human_loading(self, x):
        file = sio.loadmat(MatlabFolder('human', x.size, x.shape, x.free) + path.sep + self.VideoChain[0])
        matlab_cell = file['hats']
        for i, Frame in enumerate(matlab_cell):
            data = Frame[0]

            # to sort the data
            humans_frame = Humans_Frame(self.size)

            # if identities are given
            if data.shape[1] > 4:
                data = data[data[:, 4].argsort()]

            # are multiple hats detected?
            if data[:, 2:4].shape[0] > 1:
                humans_frame.position[self.occupied] = data[:, 2:4][0] + [x.x_error[0], x.y_error[0]]
            else:
                humans_frame.position[self.occupied] = data[:, 2:4] + [x.x_error[0], x.y_error[0]]

            # tracked participants have a carrying boolean, and an angle to their force meter
            if data.shape[1] > 5:
                humans_frame.carrying[self.occupied] = data[:, 5]
                humans_frame.angle[self.occupied] = data[:, 6] * np.pi / 180 + x.angle_error[0]

            self.frames.append(humans_frame)
        return

    def forces_loading(self, x):
        # read force meter file
        with open(force_directory(self) + path.sep + force_filename(self), 'r') as f:
            reader = csv.reader(f, delimiter='\t')
            text_file_content = [line for line in reader]

        sampled_frames = convert_to_frames(x.fps, text_file_content[1::2][1:-1])

        # load forces and set them relative to the baseline
        forces = [[float(fu) for fu in fo[0].split(' ') if len(fu) > 1] for fo in text_file_content[0::2][:-1]]
        forces = relative_to_minimum(forces)

        # all unoccupied force meters should have zero force
        empty_indices = [i for i in range(len(list(self.gender()))) if i not in self.occupied]
        for empty_index in empty_indices:
            forces[:, empty_index] = np.zeros(forces[:, empty_index].shape)
        forces_all_frames = []

        # every frame of the movie gets a force for every force meter
        for frames_index in range(len(sampled_frames) - 1):
            for ii in range(sampled_frames[frames_index], sampled_frames[frames_index + 1]):
                forces_all_frames.append(forces[frames_index])

        # find the offset of the first frame of the movie to the start of the force meter measurement
        synch_offset = synchronization_offset(self, x)

        # write the force into the self.frame[:].forces variable
        for i, force_index in enumerate(range(synch_offset, len(self.frames) + synch_offset)):
            if force_index > len(forces_all_frames) - 1:
                self.frames[i].forces = [np.nan for i in range(9)]
            else:
                self.frames[i].forces = forces_all_frames[force_index]

    def averageCarrierNumber(self):
        return self.number

    def gender(self):
        index = excel_worksheet_index(self.filename)
        gender_string = list(sheet.cell(row=index, column=17).value)
        if len(gender_string) > 1 and self.size == 'Medium':
            right_way_around = gender_string[1:]
            right_way_around.reverse()
            gender_string = gender_string[0:1] + right_way_around

        if len(gender_string) != participant_number[self.size]\
                and not (len(gender_string) in [1, 2] and self.size == 'Medium'):
            print('you have an incorrect gender string in ' + str(index))
        return {i: letter for i, letter in enumerate(gender_string) if letter != '0'}


def human_participants(x, my_load):
    return [force_attachment_positions(my_load, x)[name] for name in x.participants.occupied]


def excel_worksheet_index(filename):
    name_list = filename.split('_')
    indices = []
    for i in range(2, number_exp + 1):
        if len([ii for ii in range(len(name_list))
                if i <= number_exp and
                   name_list[ii] in sheet.cell(row=i, column=1).value.split('_')]) > 1:
            indices.append(i)
    if len(indices) == 1:
        return indices[0]
    elif len(name_list[-1]) > 1:  # has to be the first run
        return indices[int(np.argmin([sheet.cell(row=index, column=6).value for index in indices]))]
    elif len(name_list[-1]) == 1:
        return indices[np.argsort([sheet.cell(row=index, column=6).value
                                   for index in indices])[int(name_list[-1]) - 1]]
    elif len(indices) == 0:
        print('cant find your movie')
